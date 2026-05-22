from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from typing import Optional
from ..models.alarm_record import AlarmRecord
from ..collector.alarm_mapping import get_mapping


def _current_msg(code: str, plc_id: int = None) -> str:
    mapping = get_mapping(plc_id)
    for (_byte, _bit), (c, msg) in mapping.items():
        if c == code:
            return msg
    return code


def _apply_plc_filter(query, plc_id: Optional[int]):
    if plc_id is not None:
        return query.filter(AlarmRecord.plc_id == plc_id)
    return query


def get_overview(db: Session, plc_id: Optional[int] = None):
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    base = db.query(AlarmRecord)
    base = _apply_plc_filter(base, plc_id)

    total = base.count()
    active = base.filter(AlarmRecord.is_active == True).count()
    today = base.filter(AlarmRecord.start_time >= today_start).count()
    today_resolved = base.filter(AlarmRecord.end_time >= today_start).count()

    avg_dur = db.query(func.avg(AlarmRecord.duration_seconds)).filter(
        AlarmRecord.duration_seconds.isnot(None)
    )
    avg_dur = _apply_plc_filter(avg_dur, plc_id)
    avg_dur = avg_dur.scalar()

    return {
        "total_alarms": total,
        "active_alarms": active,
        "today_alarms": today,
        "today_resolved": today_resolved,
        "avg_duration_seconds": round(float(avg_dur), 1) if avg_dur else None,
    }


def get_alarm_distribution(db: Session, plc_id: Optional[int] = None):
    query = db.query(
        AlarmRecord.alarm_code,
        func.count(AlarmRecord.id).label("count"),
    )
    query = _apply_plc_filter(query, plc_id)
    results = query.group_by(AlarmRecord.alarm_code).order_by(
        func.count(AlarmRecord.id).desc()
    ).all()
    return [
        {"alarm_code": r.alarm_code, "alarm_message": _current_msg(r.alarm_code, plc_id), "count": r.count}
        for r in results
    ]


def get_daily_trend(db: Session, days: int = 7, plc_id: Optional[int] = None):
    since = datetime.utcnow() - timedelta(days=days)
    query = db.query(
        func.date(AlarmRecord.start_time).label("date"),
        func.count(AlarmRecord.id).label("count"),
    ).filter(AlarmRecord.start_time >= since)
    query = _apply_plc_filter(query, plc_id)
    results = query.group_by(func.date(AlarmRecord.start_time)).order_by("date").all()
    return [{"date": str(r.date), "count": r.count} for r in results]


def get_weekly_report(db: Session, week_start: str, plc_id: Optional[int] = None):
    week_start_dt = datetime.fromisoformat(week_start)
    week_end_dt = week_start_dt + timedelta(days=7)
    last_week_start = week_start_dt - timedelta(days=7)
    last_week_end = week_start_dt

    this_week = db.query(AlarmRecord.alarm_code,
                         func.count(AlarmRecord.id).label("count"))
    this_week = this_week.filter(AlarmRecord.start_time >= week_start_dt,
                                 AlarmRecord.start_time < week_end_dt)
    this_week = _apply_plc_filter(this_week, plc_id)
    this_week = this_week.group_by(AlarmRecord.alarm_code).all()

    last_week = db.query(AlarmRecord.alarm_code,
                         func.count(AlarmRecord.id).label("count"))
    last_week = last_week.filter(AlarmRecord.start_time >= last_week_start,
                                 AlarmRecord.start_time < last_week_end)
    last_week = _apply_plc_filter(last_week, plc_id)
    last_week = last_week.group_by(AlarmRecord.alarm_code).all()

    this_map = {r.alarm_code: (_current_msg(r.alarm_code, plc_id), r.count) for r in this_week}
    last_map = {r.alarm_code: r.count for r in last_week}

    all_codes = set(list(this_map.keys()) + list(last_map.keys()))
    items = []
    for code in all_codes:
        msg, this_cnt = this_map.get(code, (code, 0))
        last_cnt = last_map.get(code, 0)
        change = this_cnt - last_cnt
        pct = round((change / last_cnt) * 100, 1) if last_cnt > 0 else (100.0 if this_cnt > 0 else 0.0)
        items.append({
            "alarm_code": code,
            "alarm_message": msg,
            "this_week_count": this_cnt,
            "last_week_count": last_cnt,
            "change_percent": pct,
        })
    items.sort(key=lambda x: x["this_week_count"], reverse=True)
    return {"top_alarms": items}


def get_duration_stats(db: Session, week_start: str, plc_id: Optional[int] = None):
    week_start_dt = datetime.fromisoformat(week_start)
    week_end_dt = week_start_dt + timedelta(days=7)

    query = db.query(AlarmRecord.alarm_code,
                     func.sum(AlarmRecord.duration_seconds).label("total_dur"))
    query = query.filter(AlarmRecord.start_time >= week_start_dt,
                         AlarmRecord.start_time < week_end_dt,
                         AlarmRecord.duration_seconds.isnot(None))
    query = _apply_plc_filter(query, plc_id)
    results = query.group_by(AlarmRecord.alarm_code).all()
    return [{"alarm_code": r.alarm_code, "alarm_message": _current_msg(r.alarm_code, plc_id),
             "total_duration_sec": int(r.total_dur or 0)} for r in results]


def export_weekly_excel(db: Session, week_start: str, plc_id: Optional[int] = None, sort_by: str = "count"):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    week_start_dt = datetime.fromisoformat(week_start)
    week_end_dt = week_start_dt + timedelta(days=7)
    last_week_start = week_start_dt - timedelta(days=7)
    last_week_end = week_start_dt
    week_num = week_start_dt.isocalendar()[1]

    plc_name = "全部线体"
    if plc_id is not None:
        from ..models.plc_config import PLCConfig
        plc = db.query(PLCConfig).filter(PLCConfig.id == plc_id).first()
        if plc:
            plc_name = plc.name

    this_week = db.query(AlarmRecord.alarm_code,
                         func.count(AlarmRecord.id).label("cnt"),
                         func.sum(AlarmRecord.duration_seconds).label("dur"),
                         func.max(AlarmRecord.alarm_message).label("msg"))
    this_week = this_week.filter(AlarmRecord.start_time >= week_start_dt,
                                 AlarmRecord.start_time < week_end_dt)
    this_week = _apply_plc_filter(this_week, plc_id)
    this_week = this_week.group_by(AlarmRecord.alarm_code).all()

    last_week = db.query(AlarmRecord.alarm_code,
                         func.count(AlarmRecord.id).label("cnt"),
                         func.sum(AlarmRecord.duration_seconds).label("dur"))
    last_week = last_week.filter(AlarmRecord.start_time >= last_week_start,
                                 AlarmRecord.start_time < last_week_end)
    last_week = _apply_plc_filter(last_week, plc_id)
    last_week = last_week.group_by(AlarmRecord.alarm_code).all()

    def _desc(code, db_msg=None):
        m = _current_msg(code, plc_id)
        if m != code:
            return m
        if db_msg and db_msg != code:
            return db_msg
        return code

    this_map = {r.alarm_code: (_desc(r.alarm_code, r.msg), r.cnt, int(r.dur or 0)) for r in this_week}
    last_map = {r.alarm_code: (r.cnt, int(r.dur or 0)) for r in last_week}

    def fmt_dur(sec):
        if sec <= 0:
            return "0s"
        h, m = divmod(sec, 3600)
        m, s = divmod(m, 60)
        parts = []
        if h: parts.append(f"{h}h")
        if m: parts.append(f"{m}m")
        if s or not parts: parts.append(f"{s}s")
        return "".join(parts)

    wb = Workbook()
    ws = wb.active
    ws.title = "TOP10故障分析"

    # Colors
    title_fill = PatternFill(start_color="C41230", end_color="C41230", fill_type="solid")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_font = Font(color="1565C0", bold=True)
    red_font = Font(color="C41230")
    green_font = Font(color="2E7D32")

    # Title row
    sort_label = "次数" if sort_by == "count" else "时长"
    title = f"{plc_name}-W{week_num} TOP10故障分析（{sort_label}）"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border
    for c in range(2, 14):
        ws.cell(row=1, column=c).fill = title_fill
        ws.cell(row=1, column=c).border = thin_border
    ws.row_dimensions[1].height = 36

    # Subtitle row (date range)
    sub = f"统计周期: {week_start_dt.strftime('%Y-%m-%d')} ~ {(week_end_dt - timedelta(days=1)).strftime('%Y-%m-%d')}    对比周期: {last_week_start.strftime('%Y-%m-%d')} ~ {(last_week_end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    sub_cell = ws.cell(row=2, column=1, value=sub)
    sub_cell.font = Font(size=10, color="666666")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    sub_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    sub_cell.border = thin_border
    for c in range(2, 14):
        ws.cell(row=2, column=c).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws.cell(row=2, column=c).border = thin_border
    ws.row_dimensions[2].height = 24

    # Headers (row 3)
    headers = ["报警代码", "报警描述", "本周次数", "上周次数", "次数变化%",
               "本周时长", "上周时长", "时长变化%", "状态",
               "故障原因", "责任人", "完成时间", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # Data
    all_codes = set(list(this_map.keys()) + list(last_map.keys()))
    rows = []
    for code in all_codes:
        msg, this_cnt, this_dur = this_map.get(code, (code, 0, 0))
        last_cnt, last_dur = last_map.get(code, (0, 0))
        cnt_pct = round(((this_cnt - last_cnt) / last_cnt) * 100, 1) if last_cnt > 0 else (100.0 if this_cnt > 0 else 0.0)
        dur_pct = round(((this_dur - last_dur) / last_dur) * 100, 1) if last_dur > 0 else (100.0 if this_dur > 0 else 0.0)

        if this_cnt == 0 and last_cnt > 0:
            status = "已解决"
        elif last_cnt > 0 and this_cnt < last_cnt * 0.5:
            status = "大幅改善"
        elif this_cnt > 0 and last_cnt == 0:
            status = "新增"
        else:
            status = ""

        rows.append([code, msg, this_cnt, last_cnt, cnt_pct,
                     fmt_dur(this_dur), fmt_dur(last_dur), dur_pct, status,
                     "", "", "", ""])

    sort_idx = 2 if sort_by == "count" else 5
    rows.sort(key=lambda r: r[sort_idx] if isinstance(r[sort_idx], (int, float)) else 0, reverse=True)
    rows = rows[:10]

    for i, row in enumerate(rows, 4):
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            if col in (1, 3, 4, 5, 8, 9):
                cell.alignment = cell_align
            elif col in (2, 10, 11, 12, 13):
                cell.alignment = left_align
            else:
                cell.alignment = cell_align

        # Row coloring by status
        if row[8] == "已解决":
            for col in range(1, 14):
                ws.cell(row=i, column=col).fill = green_fill
        elif row[8] == "大幅改善":
            for col in range(1, 14):
                ws.cell(row=i, column=col).fill = yellow_fill
        elif row[8] == "新增":
            for col in range(1, 14):
                ws.cell(row=i, column=col).fill = red_fill

        # Color the change% cells
        cnt_pct_cell = ws.cell(row=i, column=5)
        if isinstance(row[4], (int, float)) and row[4] > 0:
            cnt_pct_cell.font = Font(color="C41230", bold=True)
        elif isinstance(row[4], (int, float)) and row[4] < 0:
            cnt_pct_cell.font = Font(color="2E7D32", bold=True)

        dur_pct_cell = ws.cell(row=i, column=8)
        if isinstance(row[7], (int, float)) and row[7] > 0:
            dur_pct_cell.font = Font(color="C41230", bold=True)
        elif isinstance(row[7], (int, float)) and row[7] < 0:
            dur_pct_cell.font = Font(color="2E7D32", bold=True)

    # Column widths
    col_widths = {'A': 10, 'B': 22, 'C': 10, 'D': 10, 'E': 12, 'F': 12,
                  'G': 12, 'H': 12, 'I': 11, 'J': 22, 'K': 12, 'L': 14, 'M': 22}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # Row heights for data
    for i in range(4, 4 + len(rows)):
        ws.row_dimensions[i].height = 22

    # Freeze header rows
    ws.freeze_panes = "A4"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
