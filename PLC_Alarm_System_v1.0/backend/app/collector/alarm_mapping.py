import os
from typing import Dict, Tuple, Optional

_mappings: Dict[int, Dict[Tuple[int, int], Tuple[str, str]]] = {}

BASE_DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'plc_data'))


def _get_default_dict_path(plc_id: int = None) -> str:
    if plc_id:
        return os.path.join(BASE_DATA_DIR, f"PLC-{plc_id}", "fault_dict.xlsx")
    return os.path.join(BASE_DATA_DIR, "default", "fault_dict.xlsx")


def _parse_excel(excel_path: str) -> Dict[Tuple[int, int], Tuple[str, str]]:
    from openpyxl import load_workbook

    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"故障字典文件不存在: {excel_path}")

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['故障字典']
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None or row[2] is None:
            continue
        byte = int(row[1])
        bit = int(row[2])
        code = str(row[3]).strip() if row[3] else ''
        msg = str(row[4]).strip() if row[4] else ''
        if code:
            mapping[(byte, bit)] = (code, msg)
    wb.close()
    return mapping


def load_mapping(plc_id: int, path: str = None):
    if path is None:
        path = _get_default_dict_path(plc_id)
    if path.endswith('.db'):
        mapping = _parse_db_file(path)
    else:
        mapping = _parse_excel(path)
    _mappings[plc_id] = mapping
    return mapping


def _parse_db_file(db_path: str) -> Dict[Tuple[int, int], Tuple[str, str]]:
    from .tia_parser import TiaDbParser
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"DB文件不存在: {db_path}")
    parser = TiaDbParser(db_path)
    result = parser.to_fault_mapping()
    if not result:
        raise FileNotFoundError(f"DB文件中未找到Bool数组定义: {db_path}")
    return result


def reload_mapping(plc_id: int) -> Dict[Tuple[int, int], Tuple[str, str]]:
    excel_path = None
    from ..core.database import SessionLocal
    from ..models.plc_config import PLCConfig
    db = SessionLocal()
    try:
        plc = db.query(PLCConfig).filter(PLCConfig.id == plc_id).first()
        if plc and plc.dict_path:
            excel_path = plc.dict_path
    finally:
        db.close()

    if not excel_path:
        excel_path = _get_default_dict_path(plc_id)
    return load_mapping(plc_id, excel_path)


def get_mapping(plc_id: Optional[int] = None) -> Dict[Tuple[int, int], Tuple[str, str]]:
    if plc_id is not None:
        if plc_id not in _mappings:
            path = _get_default_dict_path(plc_id)
            from ..core.database import SessionLocal
            from ..models.plc_config import PLCConfig
            db = SessionLocal()
            try:
                plc = db.query(PLCConfig).filter(PLCConfig.id == plc_id).first()
                if plc and plc.dict_path:
                    path = plc.dict_path
            finally:
                db.close()
            try:
                load_mapping(plc_id, path)
            except FileNotFoundError:
                _mappings[plc_id] = {}
        return _mappings[plc_id]

    if not _mappings:
        from ..core.database import SessionLocal
        from ..models.plc_config import PLCConfig
        db = SessionLocal()
        try:
            plcs = db.query(PLCConfig).filter(PLCConfig.is_active == True).all()
            for plc in plcs:
                if plc.id not in _mappings:
                    path = plc.dict_path or _get_default_dict_path(plc.id)
                    try:
                        if path.endswith('.db'):
                            _mappings[plc.id] = _parse_db_file(path)
                        else:
                            _mappings[plc.id] = _parse_excel(path)
                    except FileNotFoundError:
                        pass
            if not plcs:
                load_mapping(0)
        finally:
            db.close()

    merged = {}
    for m in _mappings.values():
        merged.update(m)
    if merged:
        return merged
    return {}


ALARM_MAPPING = _mappings
