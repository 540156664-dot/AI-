import os, re, json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

base = r"d:\360MoveData\Users\Administrator\Desktop\车端监控程序\CAN诊断分析工具\用户资料\actuator_msgs\actuator_msgs\msg"
cfg_base = r"d:\360MoveData\Users\Administrator\Desktop\车端监控程序\CAN诊断分析工具\用户资料"

msg_files = [
    "actuator_cmd_msg.msg", "actuator_general_states.msg", "actuator_msg.msg",
    "actuator_string.msg", "battery_state_msg.msg", "Float64MapArray.msg",
    "fork_sensor_msg.msg", "front_safe_msg.msg", "io.msg",
    "ManualCalibrate.msg", "ManualCalibrateFeedback.msg", "ManualCalibrateField.msg",
    "rail_sensors.msg", "safe_sensor_msg.msg", "servo_state.msg",
    "steer_state_msg.msg", "sub_steer_state_msg.msg", "vehicle_state_msg.msg",
    "WorkStateMsg.msg", "collision_sensor.msg", "collision_sensor_all.msg",
    "collision_unlock.msg", "collision_unlock_all.msg", "load_weight.msg",
]

PRIMITIVES = {"bool","int8","uint8","int16","uint16","int32","uint32","float32","float64","string","time","duration","Header"}
EXTERNAL = {"diagnose_msgs/Float64Map","diagnose_msgs/StringMap","diagnose_msgs/BoolMap","diagnose_msgs/Int32Map"}

def parse_msg(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    lines = content.strip().split("\n")
    fields = []
    constants = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        cm = re.match(r'(int\d+|float\d+)\s+(\w+)\s*=\s*(-?\d+)', line)
        if cm:
            constants.append((cm.group(2), cm.group(3)))
            continue
        m = re.match(r'(\S+(?:/\S+)?)\s+(\w+)\s*(#\s*(.*))?', line)
        if m:
            ftype = m.group(1); fname = m.group(2); comment = m.group(4) or ""
            fields.append((ftype, fname, comment))
    fields = [(t,n,c) for t,n,c in fields if n not in ("header","head")]
    return fields, constants

type_registry = {}
all_constants = {}
for fname in msg_files:
    fields, constants = parse_msg(os.path.join(base, fname))
    type_registry[fname.replace(".msg","")] = fields
    if constants:
        all_constants[fname.replace(".msg","")] = constants

def flatten_fields(fields, prefix="", depth=0, max_depth=6):
    result = []
    for ftype, fname_field, comment in fields:
        is_array = "[]" in ftype
        base_type = ftype.replace("[]","")
        full_name = f"{prefix}{fname_field}"
        is_custom = base_type in type_registry and base_type not in PRIMITIVES and base_type not in EXTERNAL and not base_type.startswith("diagnose_msgs/")
        if is_custom and depth < max_depth:
            sub_fields = type_registry[base_type]
            sub_prefix = f"{full_name}." if not is_array else f"{full_name}[i]."
            for sftype, sfname, sfcomment in sub_fields:
                sisarr = "[]" in sftype
                sbase = sftype.replace("[]","")
                siscust = sbase in type_registry and sbase not in PRIMITIVES and sbase not in EXTERNAL and not sbase.startswith("diagnose_msgs/")
                if siscust and depth+1 < max_depth:
                    inner = flatten_fields([(sftype, sfname, sfcomment)], prefix=sub_prefix, depth=depth+1, max_depth=max_depth)
                    result.extend(inner)
                else:
                    inner_name = f"{sub_prefix}{sfname}"
                    if sisarr: inner_name += "[j]"
                    result.append((sbase, inner_name, sfcomment))
        else:
            result.append((base_type, full_name, comment))
    return result

# Load CAN config tags
with open(os.path.join(cfg_base, "DeviceCanModelObjectConfig.json"), "r", encoding="utf-8") as f:
    dev_config = json.load(f)

# Extract all tagName -> tagDescription mappings
tag_map = {}
for device in dev_config.get("deviceModelObjectTemplate", []):
    className = device.get("className","")
    for tag in device.get("deviceComponentObjectTagInfos", []):
        tname = tag.get("tagName","")
        tdesc = tag.get("tagDescription","")
        canID = tag.get("canID", 0)
        ttype = tag.get("tagDataType", 0)
        if tname:
            if tname not in tag_map:
                tag_map[tname] = []
            tag_map[tname].append({
                "className": className,
                "description": tdesc,
                "canID": canID,
                "dataType": ttype
            })

# dataType map: 1=bool, 2=uint8, 3=uint16, 4=uint32, 6=int16/int32, 7=float32, 10=float64 (approx)
type_name = {1:"bool", 2:"uint8", 3:"uint16", 4:"uint32", 6:"int16/int32", 7:"float32", 10:"float64"}

# Manual semantic mapping: ROS field -> CAN tagName (best guess from context)
semantic_map = {
    "walking_speed": "speed_vcu",
    "steer_angle": "fb_steer",
    "angle_speed": None,
    "is_auto_mode": "autoState",
    "is_communicate_normal": None,
    "unbrake_enable": None,
    "steer_state_msg.steer_speed": "steering_command",
    "steer_state_msg.steer_angle": "SteerAngle",
    "steer_state_msg.error_code": "fb_steer_1_error_code",
    "sub_steer_state_msg.steer_speed": None,
    "sub_steer_state_msg.steer_angle": None,
    "sub_steer_state_msg.error_code": None,
    "servo_state.servo_current": None,
    "servo_state.servo_torque": None,
    "battery_state_msg.voltage_soc": "SOC",
    "battery_state_msg.voltage_tem": "temperature",
    "battery_state_msg.voltage_vol": "voltage",
    "battery_state_msg.voltage_cul": "BatteryCurrent",
    "battery_state_msg.is_charge": "chargeState",
    "battery_state_msg.warning": "faultState",
    "WorkStateMsg.type": None,
    "WorkStateMsg.id": None,
    "WorkStateMsg.pose": "cableEncoderLength",
    "WorkStateMsg.speed": "cableEncoderSpeed",
    "WorkStateMsg.is_load": None,
    "safe_sensor_msg.trigger_area": None,
    "front_safe_msg.trigger_area": None,
    "fork_sensor_msg.is_trigger": "leftCargoDetection",
    "collision_sensor.type": None,
    "collision_sensor.id": None,
    "collision_sensor.is_trigger": "leftBumpProtection",
    "collision_sensor.distance": None,
    "collision_unlock.type": None,
    "collision_unlock.id": None,
    "collision_unlock.signal": None,
    "load_weight.total_weight": None,
    "load_weight.net_weight": None,
    "load_weight.load_state": None,
    "vehicle_state_msg.speed": "speed_vcu",
    "vehicle_state_msg.angle_speed": None,
    "rail_sensors.sensor_install": None,
    "rail_sensors.rfid": None,
    "rail_sensors.is_flipped": None,
    "actuator_cmd_msg.lift_height": None,
    "actuator_cmd_msg.light": "redLight",
    "actuator_cmd_msg.sound": "volume",
    "actuator_cmd_msg.safe_area": None,
    "actuator_cmd_msg.actuator_mode": None,
}

# ===== Build Excel =====
wb = openpyxl.Workbook()

hfont = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
hfill2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
msg_font = Font(name="Microsoft YaHei", bold=True, size=11, color="1F4E79")
msg_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ff = Font(name="Consolas", size=10)
cf = Font(name="Microsoft YaHei", size=9, color="666666")
warn_f = Font(name="Microsoft YaHei", size=9, color="CC0000")
match_f = Font(name="Consolas", size=10, color="006600")
nomatch_f = Font(name="Consolas", size=9, color="999999")
thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_header(ws, row, headers, widths, fill=hfill):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hfont; c.fill = fill; c.alignment = ca; c.border = thin
        ws.column_dimensions[get_column_letter(col)].width = w

# ====== Sheet 1: 原始定义 ======
ws1 = wb.active
ws1.title = "原始定义"
set_header(ws1, 1, ["消息文件","字段序号","字段类型","字段名称","数组","注释说明"], [28,8,26,30,6,55])
ws1.freeze_panes = "A2"
r = 2
for fname in msg_files:
    fields, constants = parse_msg(os.path.join(base, fname))
    mb = fname.replace(".msg","")
    ws1.cell(row=r, column=1, value=mb).font = msg_font
    ws1.cell(row=r, column=1).fill = msg_fill
    ws1.cell(row=r, column=1).alignment = la
    ws1.cell(row=r, column=1).border = thin
    for c in range(2,7): ws1.cell(row=r, column=c).fill = msg_fill; ws1.cell(row=r, column=c).border = thin
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    if mb in all_constants:
        for cn, cv in all_constants[mb]:
            for c in range(1,7): ws1.cell(row=r, column=c).border = thin
            ws1.cell(row=r, column=3, value="常量").font = Font(name="Consolas", size=10, color="996633")
            ws1.cell(row=r, column=4, value=f"{cn} = {cv}").font = ff; ws1.cell(row=r, column=4).alignment = la
            r += 1
    for idx, (ft,fname_field,comment) in enumerate(fields,1):
        is_arr = "[]" in ft; dt = ft.replace("[]","")
        for c in range(1,7): ws1.cell(row=r, column=c).border = thin
        ws1.cell(row=r, column=2, value=idx).font = ff; ws1.cell(row=r, column=2).alignment = ca
        ws1.cell(row=r, column=3, value=dt).font = ff
        ws1.cell(row=r, column=4, value=fname_field).font = ff
        ws1.cell(row=r, column=5, value="[]" if is_arr else "").font = ff; ws1.cell(row=r, column=5).alignment = ca
        ws1.cell(row=r, column=6, value=comment).font = cf
        for c in [1,3,4,6]: ws1.cell(row=r, column=c).alignment = la
        r += 1
    r += 1

# ====== Sheet 2: 展开视图 ======
ws2 = wb.create_sheet("展开视图")
set_header(ws2, 1, ["消息文件","字段序号","底层类型","字段路径（完全展开）","注释说明"], [28,8,24,55,50])
ws2.freeze_panes = "A2"
r2 = 2
for fname in msg_files:
    mb = fname.replace(".msg","")
    fpath = os.path.join(base, fname)
    fields, constants = parse_msg(fpath)
    flat = flatten_fields(fields)
    ws2.cell(row=r2, column=1, value=mb).font = msg_font
    ws2.cell(row=r2, column=1).fill = msg_fill
    ws2.cell(row=r2, column=1).alignment = la
    ws2.cell(row=r2, column=1).border = thin
    for c in range(2,6): ws2.cell(row=r2, column=c).fill = msg_fill; ws2.cell(row=r2, column=c).border = thin
    ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=5)
    r2 += 1
    if mb in all_constants:
        for cn, cv in all_constants[mb]:
            for c in range(1,6): ws2.cell(row=r2, column=c).border = thin
            ws2.cell(row=r2, column=3, value="常量").font = Font(name="Consolas", size=10, color="996633")
            ws2.cell(row=r2, column=4, value=f"{cn} = {cv}").font = ff; ws2.cell(row=r2, column=4).alignment = la
            r2 += 1
    for idx, (ft, path, comment) in enumerate(flat, 1):
        for c in range(1,6): ws2.cell(row=r2, column=c).border = thin
        ws2.cell(row=r2, column=2, value=idx).font = ff; ws2.cell(row=r2, column=2).alignment = ca
        ws2.cell(row=r2, column=3, value=ft).font = ff
        ws2.cell(row=r2, column=4, value=path).font = ff
        ws2.cell(row=r2, column=5, value=comment).font = cf
        for c in [1,3,4,5]: ws2.cell(row=r2, column=c).alignment = la
        is_ext = ft in EXTERNAL or ft.replace("[]","") in EXTERNAL
        if is_ext: ws2.cell(row=r2, column=3).font = Font(name="Consolas", size=10, color="996633")
        r2 += 1
    r2 += 1

# ====== Sheet 3: 自定义类型字典 ======
ws3 = wb.create_sheet("自定义类型字典")
set_header(ws3, 1, ["自定义类型","原始类型字段列表","说明"], [26,52,48])
ws3.freeze_panes = "A2"
cti = {
    "steer_state_msg": ("float32 steer_speed\nfloat32 steer_angle\nint32 error_code\nsub_steer_state_msg[] sub_steers_state","舵轮转向状态（含子舵轮列表）"),
    "sub_steer_state_msg": ("float32 steer_speed\nfloat32 steer_angle\nint32 error_code","子舵轮转向状态"),
    "servo_state": ("float32 servo_current\nfloat32 servo_torque","伺服驱动器电流/扭矩"),
    "battery_state_msg": ("uint8 voltage_soc\nuint16 voltage_tem\nuint16 voltage_vol\nuint16 voltage_cul\nuint8 is_charge\nuint8 warning","电池状态(SOC/温度/电压/电流/充电/告警)"),
    "WorkStateMsg": ("uint8 type\nuint8 id\nfloat32 pose\nfloat32 speed\nbool is_load","上装工作状态(举升/伸叉/旋转等)"),
    "safe_sensor_msg": ("uint8 trigger_area","安全传感器(0-无 1-减速 2-停止)"),
    "front_safe_msg": ("uint8 trigger_area","前向安全传感器"),
    "fork_sensor_msg": ("bool is_trigger","货叉传感器触发"),
    "collision_sensor": ("time stamp\nuint8 type\nuint8 id\nbool is_trigger\nfloat32 distance\nstring[] vec_str","碰撞传感器(0未知 1光电 2机械 3测距)"),
    "collision_unlock": ("uint8 type\nuint8 id\nuint8 signal","碰撞解锁信号"),
    "load_weight": ("uint8 version\nint8 load_state\nfloat32 total_weight\nfloat32 net_weight\nfloat32 uncertainty\nfloat32 thresh_under\nfloat32 thresh_zero\nfloat32 thresh_over\nstring extra_attr\n常量:Invalid=-3 Ambiguous=-2 Underload=-1 Zero=0 NormalLoad=1 Overload=2","称重数据(含负载枚举和阈值)"),
    "ManualCalibrateField": ("int32 iWheeId\nfloat32 fWheelRaidus\nfloat32 fSteerAngleOffset","标定字段(轮ID/半径/转角偏移)"),
    "diagnose_msgs/Float64Map": ("string key\nfloat64 value","外部:浮点键值对"),
    "diagnose_msgs/StringMap": ("string key\nstring value","外部:字符串键值对"),
    "diagnose_msgs/BoolMap": ("string key\nbool value","外部:布尔键值对"),
    "diagnose_msgs/Int32Map": ("string key\nint32 value","外部:整型键值对"),
}
r3 = 2
for tname, (fs, desc) in cti.items():
    for col, val in enumerate([tname, fs, desc], 1):
        c = ws3.cell(row=r3, column=col, value=val)
        c.font = Font(name="Consolas", size=10) if col<=2 else Font(name="Microsoft YaHei", size=10)
        c.alignment = la; c.border = thin
    ws3.cell(row=r3, column=1).font = Font(name="Consolas", size=10, bold=True, color="1F4E79")
    nlines = fs.count("\n")+1
    ws3.row_dimensions[r3].height = max(15, 15*nlines)
    r3 += 1

# ====== Sheet 4: TAG映射 ======
ws4 = wb.create_sheet("TAG映射（ROS字段↔CAN配置）")
set_header(ws4, 1, ["ROS消息", "ROS字段路径", "ROS类型", "匹配CAN tagName", "CAN描述 (tagDescription)", "设备类", "匹配方式", "备注"], [20,42,16,24,28,12,10,30])
ws4.freeze_panes = "A2"

# Build all ROS fields from main messages (actuator_msg, actuator_cmd_msg, battery_state_msg, etc.)
main_msgs = ["actuator_msg", "actuator_cmd_msg"]
all_ros_fields = []
for mn in main_msgs:
    fields = type_registry.get(mn, [])
    flat = flatten_fields(fields, prefix=f"{mn}.")
    for ft, path, comment in flat:
        all_ros_fields.append((mn, path, ft, comment))

r4 = 2
for mn, path, ft, comment in all_ros_fields:
    # Try to match
    # Strategy: look at the last part of the path, try fuzzy match with tag_map keys
    leaf = path.rsplit(".", 1)[-1].replace("[i]","").replace("[j]","")

    # Direct match
    matched = tag_map.get(leaf, [])
    match_type = ""
    can_tag = ""
    can_desc = ""
    can_class = ""
    note = ""

    if matched:
        # Take first match
        m = matched[0]
        can_tag = leaf
        can_desc = m["description"]
        can_class = m["className"]
        match_type = "名称匹配"
    else:
        # Try semantic map
        # Build key from parent type + field
        for pattern, suggested_tag in semantic_map.items():
            if path.endswith(pattern) or leaf == pattern.split(".")[-1]:
                if suggested_tag and suggested_tag in tag_map:
                    m = tag_map[suggested_tag][0]
                    can_tag = suggested_tag
                    can_desc = m["description"]
                    can_class = m["className"]
                    match_type = "语义推断"
                    note = "需人工确认"
                    break
                elif suggested_tag:
                    can_tag = suggested_tag
                    can_desc = "(tag未在配置中找到)"
                    match_type = "语义推断"
                    note = "tag不存在，需确认"
                    break
        else:
            # Check if it's a simple leaf that might match by partial name
            if leaf in tag_map:
                m = tag_map[leaf][0]
                can_tag = leaf
                can_desc = m["description"]
                can_class = m["className"]
                match_type = "名称匹配"
            else:
                can_tag = "—"
                can_desc = "—"
                can_class = "—"
                match_type = "无法匹配"
                note = "需手动配置映射"

    # Highlight unmatched
    row_fill = None
    if match_type == "无法匹配":
        row_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for col, val in enumerate([mn, path, ft, can_tag, can_desc, can_class, match_type, note], 1):
        c = ws4.cell(row=r4, column=col, value=val)
        if col == 2: c.font = Font(name="Consolas", size=9)
        elif col == 4: c.font = match_f if match_type != "无法匹配" else nomatch_f
        elif col == 7: c.font = Font(name="Microsoft YaHei", size=9, color="006600" if match_type != "无法匹配" else "CC0000")
        else: c.font = Font(name="Microsoft YaHei", size=9)
        c.alignment = la if col in (2,5,8) else ca
        c.border = thin
        if row_fill: c.fill = row_fill
    r4 += 1

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 44
ws4.column_dimensions['C'].width = 14
ws4.column_dimensions['D'].width = 26
ws4.column_dimensions['E'].width = 30
ws4.column_dimensions['F'].width = 12
ws4.column_dimensions['G'].width = 10
ws4.column_dimensions['H'].width = 32

# ====== Sheet 5: CAN配置Tag速查 ======
ws5 = wb.create_sheet("CAN配置Tag速查")
set_header(ws5, 1, ["tagName","tagDescription","设备类","CAN-ID","数据类型","信号偏移(tagOffsetAddress)"], [24,30,16,10,12,22])
ws5.freeze_panes = "A2"
r5 = 2
seen_tags = set()
for device in dev_config.get("deviceModelObjectTemplate", []):
    cn = device.get("className","")
    for tag in device.get("deviceComponentObjectTagInfos", []):
        tn = tag.get("tagName","")
        if tn in seen_tags: continue
        seen_tags.add(tn)
        td = tag.get("tagDescription","")
        cid = tag.get("canID",0)
        dt = tag.get("tagDataType",0)
        offset = tag.get("tagOffsetAddress","")
        dt_name = type_name.get(dt, str(dt))
        for col, val in enumerate([tn, td, cn, cid, dt_name, offset], 1):
            c = ws5.cell(row=r5, column=col, value=val)
            c.font = Font(name="Microsoft YaHei", size=9) if col in (2,3) else Font(name="Consolas", size=9)
            c.alignment = la if col in (2,3,6) else ca; c.border = thin
        r5 += 1

ws5.column_dimensions['A'].width = 24
ws5.column_dimensions['B'].width = 32
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 10
ws5.column_dimensions['E'].width = 12
ws5.column_dimensions['F'].width = 22

# ====== Sheet 6: 消息概览 ======
ws6 = wb.create_sheet("消息概览")
set_header(ws6, 1, ["序号","消息名称","消息说明","字段数","展开后字段数","依赖自定义类型"], [6,26,55,8,10,42])
ws6.freeze_panes = "A2"

summaries = {
    "actuator_cmd_msg": ("执行器控制指令","steer_state_msg"),
    "actuator_general_states": ("执行器通用状态（浮点Map数组）","diagnose_msgs/Float64Map"),
    "actuator_msg": ("执行器综合状态 — 主消息","steer_state_msg,WorkStateMsg,battery_state_msg,safe_sensor_msg"),
    "actuator_string": ("执行器字符串状态","diagnose_msgs/StringMap"),
    "battery_state_msg": ("电池状态",""),
    "Float64MapArray": ("浮点Map数组","diagnose_msgs/Float64Map"),
    "fork_sensor_msg": ("货叉传感器触发状态",""),
    "front_safe_msg": ("前向安全传感器",""),
    "io": ("IO状态（数字/整数/浮点三类Map）","diagnose_msgs/BoolMap,Int32Map,Float64Map"),
    "ManualCalibrate": ("手动标定请求",""),
    "ManualCalibrateFeedback": ("手动标定反馈","ManualCalibrateField"),
    "ManualCalibrateField": ("手动标定字段（单轮参数）",""),
    "rail_sensors": ("轨道传感器(RFID/磁条)",""),
    "safe_sensor_msg": ("安全传感器",""),
    "servo_state": ("伺服状态(电流/扭矩)",""),
    "steer_state_msg": ("舵轮转向状态","sub_steer_state_msg"),
    "sub_steer_state_msg": ("子舵轮转向状态",""),
    "vehicle_state_msg": ("车辆运动状态","servo_state"),
    "WorkStateMsg": ("上装工作状态",""),
    "collision_sensor": ("碰撞传感器(光电/机械/测距)",""),
    "collision_sensor_all": ("全部碰撞传感器集合","collision_sensor"),
    "collision_unlock": ("碰撞解锁信号",""),
    "collision_unlock_all": ("全部碰撞解锁信号集合","collision_unlock"),
    "load_weight": ("称重数据(含阈值常量)",""),
}

for i, fname in enumerate(msg_files, 1):
    mb = fname.replace(".msg","")
    desc, deps = summaries.get(mb, ("",""))
    fields, _ = parse_msg(os.path.join(base, fname))
    orig = len(fields)
    flat_len = len(flatten_fields(fields))
    for col, val in enumerate([i, mb, desc, orig, flat_len, deps], 1):
        c = ws6.cell(row=i+1, column=col, value=val)
        if col in (4,5): c.font = Font(name="Consolas", size=10); c.alignment = ca
        else: c.font = Font(name="Microsoft YaHei", size=10); c.alignment = la if col in (3,6) else ca
        c.border = thin

ws6.column_dimensions['A'].width = 6
ws6.column_dimensions['B'].width = 26
ws6.column_dimensions['C'].width = 55
ws6.column_dimensions['D'].width = 8
ws6.column_dimensions['E'].width = 10
ws6.column_dimensions['F'].width = 42

out_path = r"d:\360MoveData\Users\Administrator\Desktop\车端监控程序\actuator_msgs_definition.xlsx"
wb.save(out_path)
print(f"Done: {os.path.basename(out_path)}")

# Stats
print("\n=== 匹配统计 ===")
matched = sum(1 for mn in main_msgs for ft, path, comment in flatten_fields(type_registry[mn])
              if any(path.endswith(k) or any(path.endswith(v) for v in semantic_map.values() if v and v in tag_map) for k in tag_map)
              or (lambda leaf: leaf in tag_map)(path.rsplit(".",1)[-1].replace("[i]","").replace("[j]","")))
total = sum(1 for mn in main_msgs for _ in flatten_fields(type_registry[mn]))
print(f"  actuator_msg + actuator_cmd_msg: {total} 字段")
mapped = 0
unmapped = 0
for mn in main_msgs:
    flat = flatten_fields(type_registry[mn])
    for ft, path, comment in flat:
        leaf = path.rsplit(".",1)[-1].replace("[i]","").replace("[j]","")
        # check semantic map
        found = False
        for pat, stag in semantic_map.items():
            if path.endswith(pat) or leaf == pat.split(".")[-1]:
                if stag and stag in tag_map:
                    found = True; break
        if not found and leaf in tag_map:
            found = True
        if found: mapped += 1
        else: unmapped += 1
print(f"  可匹配: {mapped}")
print(f"  无法自动匹配: {unmapped}")
print(f"\n结论: ROS字段名和CAN tagName不是同一套命名体系，无法全自动匹配")
print(f"需要人工确认映射关系后填入 topic_subscribe.json")
